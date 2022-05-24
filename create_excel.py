import openpyxl as px

#新規作成(オブジェクト生成)から一旦保存。
wb = px.Workbook()
wb.save('test.xlsx')

#既存のファイルを読み込み
wb          = px.load_workbook('test.xlsx')

#アクティブシートを選択(新規作成時に最初からあるシート)
ws          = wb.active

#シート名を変更
ws.title    = "領収書"

#セルに値を入力していく
import datetime

ws["A1"].value  = "決済日"
ws["A2"].value  = str(datetime.date.today())

ws["B1"].value  = "商品名"
ws["C1"].value  = "個数"
ws["D1"].value  = "小計"

ws["B2"].value  = "商品A"
ws["C2"].value  = "2個"
ws["D2"].value  = "20000"

ws["B3"].value  = "商品B"
ws["C3"].value  = "1個"
ws["D3"].value  = "30000"


#セルの値を入手。計算結果を入力
ws["F2"].value  = "請求金額"
ws["F3"].value  = int(ws["D2"].value) + int(ws["D3"].value)

wb.save('test.xlsx')
